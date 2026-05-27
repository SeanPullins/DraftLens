#!/usr/bin/env python3
"""Scan the local data folder and produce a schema-agnostic inventory.

Reads are non-destructive. Profiling prefers the standard library (so it runs
with no third-party deps for CSV/JSON); Excel/Parquet/DuckDB use optional
libraries and degrade gracefully with a warning if those aren't installed.

Outputs:
  data/processed/inventory.json   machine-readable profile of every table
  docs/data_inventory.md          human-readable inventory

Usage:
  python3 scripts/inventory_data.py [--data-dir PATH] [--max-sample N]
"""

from __future__ import annotations

import argparse
import csv
import json
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import dl_common as dl


# --------------------------------------------------------------------------- #
# Type sniffing for stdlib CSV reads
# --------------------------------------------------------------------------- #
def _sniff_dtype(values: list[str]) -> str:
    seen_int = seen_float = seen_bool = seen_any = False
    for v in values:
        s = (v or "").strip()
        if s == "":
            continue
        seen_any = True
        low = s.lower()
        if low in ("true", "false"):
            seen_bool = True
            continue
        try:
            int(s)
            seen_int = True
            continue
        except ValueError:
            pass
        try:
            float(s)
            seen_float = True
            continue
        except ValueError:
            return "string"
    if not seen_any:
        return "empty"
    if seen_bool and not (seen_int or seen_float):
        return "bool"
    if seen_float:
        return "float"
    if seen_int:
        return "int"
    return "string"


def _finish(profile: dl.TableProfile, columns: list[str], dtypes: dict[str, str], rows: int) -> dl.TableProfile:
    profile.columns = columns
    profile.dtypes = dtypes
    profile.rows = rows
    profile.sample_columns = columns[:12]
    profile.purpose = dl.infer_purpose(profile.name, columns)
    profile.join_keys = sorted(
        {dl.classify_column(c) for c in columns} - {None}  # type: ignore[arg-type]
    )
    profile.outcome_fields = [c for c in columns if dl.classify_outcome(c)]
    return profile


# --------------------------------------------------------------------------- #
# Per-format profilers
# --------------------------------------------------------------------------- #
def profile_delimited(sf: dl.SourceFile, max_sample: int) -> dl.TableProfile:
    delim = "\t" if sf.ext == "tsv" else ","
    profile = dl.TableProfile(name=sf.path.name, source=str(sf.path), fmt=sf.ext, rows=0, columns=[])
    try:
        with open(sf.path, "r", encoding="utf-8-sig", newline="") as fh:
            if sf.ext == "csv":
                head = fh.read(4096)
                fh.seek(0)
                try:
                    delim = csv.Sniffer().sniff(head, delimiters=",;\t|").delimiter
                except csv.Error:
                    delim = ","
            reader = csv.reader(fh, delimiter=delim)
            header = next(reader, [])
            columns = [c.strip() for c in header]
            sample_cols: dict[str, list[str]] = {c: [] for c in columns}
            rows = 0
            for row in reader:
                rows += 1
                if rows <= max_sample:
                    for i, c in enumerate(columns):
                        sample_cols[c].append(row[i] if i < len(row) else "")
            dtypes = {c: _sniff_dtype(vals) for c, vals in sample_cols.items()}
        return _finish(profile, columns, dtypes, rows)
    except Exception as exc:  # noqa: BLE001
        profile.warnings.append(f"could not read: {exc}")
        return profile


def profile_json(sf: dl.SourceFile, max_sample: int) -> dl.TableProfile:
    profile = dl.TableProfile(name=sf.path.name, source=str(sf.path), fmt=sf.ext, rows=0, columns=[])
    try:
        records: list[dict[str, Any]] = []
        if sf.ext == "ndjson":
            with open(sf.path, "r", encoding="utf-8") as fh:
                for line in fh:
                    line = line.strip()
                    if line:
                        records.append(json.loads(line))
        else:
            with open(sf.path, "r", encoding="utf-8") as fh:
                data = json.load(fh)
            if isinstance(data, list):
                records = [r for r in data if isinstance(r, dict)]
            elif isinstance(data, dict):
                # Find the first list-of-dicts value, else treat the dict as one row.
                array = next((v for v in data.values() if isinstance(v, list) and v and isinstance(v[0], dict)), None)
                records = array if array is not None else [data]
        columns: list[str] = []
        for r in records[: max_sample or len(records)]:
            for k in r.keys():
                if k not in columns:
                    columns.append(k)
        sample = {c: [str(r.get(c, "")) for r in records[:max_sample]] for c in columns}
        dtypes = {c: _sniff_dtype(v) for c, v in sample.items()}
        return _finish(profile, columns, dtypes, len(records))
    except Exception as exc:  # noqa: BLE001
        profile.warnings.append(f"could not read: {exc}")
        return profile


def profile_excel(sf: dl.SourceFile, max_sample: int) -> dl.TableProfile:
    profile = dl.TableProfile(name=sf.path.name, source=str(sf.path), fmt=sf.ext, rows=0, columns=[])
    try:
        import openpyxl  # type: ignore
    except ImportError:
        profile.warnings.append("openpyxl not installed — run `pip install -r requirements.txt`")
        return profile
    try:
        wb = openpyxl.load_workbook(sf.path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows_iter = ws.iter_rows(values_only=True)
        header = next(rows_iter, ()) or ()
        columns = [str(c).strip() for c in header if c is not None]
        sample_cols: dict[str, list[str]] = {c: [] for c in columns}
        rows = 0
        for row in rows_iter:
            rows += 1
            if rows <= max_sample:
                for i, c in enumerate(columns):
                    sample_cols[c].append("" if i >= len(row) or row[i] is None else str(row[i]))
        wb.close()
        dtypes = {c: _sniff_dtype(v) for c, v in sample_cols.items()}
        if len(wb.sheetnames) > 1:
            profile.warnings.append(f"workbook has {len(wb.sheetnames)} sheets; only '{ws.title}' profiled")
        return _finish(profile, columns, dtypes, rows)
    except Exception as exc:  # noqa: BLE001
        profile.warnings.append(f"could not read: {exc}")
        return profile


def profile_parquet(sf: dl.SourceFile, max_sample: int) -> dl.TableProfile:
    profile = dl.TableProfile(name=sf.path.name, source=str(sf.path), fmt=sf.ext, rows=0, columns=[])
    try:
        import pyarrow.parquet as pq  # type: ignore
    except ImportError:
        profile.warnings.append("pyarrow not installed — run `pip install -r requirements.txt`")
        return profile
    try:
        pf = pq.ParquetFile(sf.path)
        schema = pf.schema_arrow
        columns = list(schema.names)
        dtypes = {name: str(schema.field(name).type) for name in columns}
        rows = pf.metadata.num_rows
        _ = max_sample
        return _finish(profile, columns, dtypes, rows)
    except Exception as exc:  # noqa: BLE001
        profile.warnings.append(f"could not read: {exc}")
        return profile


def profile_duckdb(sf: dl.SourceFile, max_sample: int) -> list[dl.TableProfile]:
    try:
        import duckdb  # type: ignore
    except ImportError:
        p = dl.TableProfile(name=sf.path.name, source=str(sf.path), fmt="duckdb", rows=0, columns=[])
        p.warnings.append("duckdb not installed — run `pip install -r requirements.txt`")
        return [p]
    profiles: list[dl.TableProfile] = []
    try:
        con = duckdb.connect(str(sf.path), read_only=True)
        tables = [r[0] for r in con.execute("SHOW TABLES").fetchall()]
        for tbl in tables:
            p = dl.TableProfile(name=f"{sf.path.name}::{tbl}", source=str(sf.path), fmt="duckdb", rows=0, columns=[])
            info = con.execute(f'PRAGMA table_info("{tbl}")').fetchall()
            columns = [row[1] for row in info]
            dtypes = {row[1]: str(row[2]) for row in info}
            rows = con.execute(f'SELECT COUNT(*) FROM "{tbl}"').fetchone()[0]
            _ = max_sample
            profiles.append(_finish(p, columns, dtypes, rows))
        con.close()
    except Exception as exc:  # noqa: BLE001
        p = dl.TableProfile(name=sf.path.name, source=str(sf.path), fmt="duckdb", rows=0, columns=[])
        p.warnings.append(f"could not read: {exc}")
        profiles.append(p)
    return profiles


def profile_file(sf: dl.SourceFile, max_sample: int) -> list[dl.TableProfile]:
    if sf.ext in ("csv", "tsv"):
        return [profile_delimited(sf, max_sample)]
    if sf.ext in ("json", "ndjson"):
        return [profile_json(sf, max_sample)]
    if sf.ext in ("xlsx", "xls"):
        return [profile_excel(sf, max_sample)]
    if sf.ext == "parquet":
        return [profile_parquet(sf, max_sample)]
    if sf.ext == "duckdb":
        return profile_duckdb(sf, max_sample)
    p = dl.TableProfile(name=sf.path.name, source=str(sf.path), fmt=sf.ext, rows=0, columns=[])
    p.warnings.append("unsupported format")
    return [p]


# --------------------------------------------------------------------------- #
# Reporting
# --------------------------------------------------------------------------- #
def render_markdown(data_dir: Path, profiles: list[dl.TableProfile], generated_at: str) -> str:
    lines = [
        "# DraftLens Data Inventory",
        "",
        f"_Generated {generated_at}_",
        "",
        f"- Source directory: `{data_dir}`",
        f"- Tables found: **{len(profiles)}**",
        "",
        "> This file is auto-generated by `scripts/inventory_data.py`. It describes",
        "> structure only (column names, counts) — no raw player values are written.",
        "",
    ]
    by_purpose: dict[str, list[dl.TableProfile]] = {}
    for p in profiles:
        by_purpose.setdefault(p.purpose, []).append(p)

    lines.append("## Summary by inferred purpose\n")
    lines.append("| Purpose | Tables | Total rows |")
    lines.append("| --- | ---: | ---: |")
    for purpose in sorted(by_purpose):
        group = by_purpose[purpose]
        lines.append(f"| {purpose} | {len(group)} | {sum(p.rows for p in group):,} |")
    lines.append("")

    lines.append("## Tables\n")
    for p in sorted(profiles, key=lambda x: (x.purpose, x.name)):
        lines.append(f"### `{p.name}`")
        lines.append("")
        lines.append(f"- Path: `{p.source}`")
        lines.append(f"- Format: {p.fmt} · Rows: {p.rows:,} · Columns: {len(p.columns)}")
        lines.append(f"- Inferred purpose: **{p.purpose}**")
        if p.join_keys:
            lines.append(f"- Candidate join keys: {', '.join(p.join_keys)}")
        if p.outcome_fields:
            lines.append(f"- Candidate outcome fields (labels only): {', '.join(p.outcome_fields)}")
        if p.sample_columns:
            lines.append(f"- Sample columns: {', '.join(p.sample_columns)}")
        if p.warnings:
            for w in p.warnings:
                lines.append(f"- ⚠️ {w}")
        lines.append("")
    return "\n".join(lines)


def main() -> int:
    cfg = dl.load_config()
    ap = argparse.ArgumentParser(description="Inventory the local DraftLens data folder.")
    ap.add_argument("--data-dir", default=None, help="Override the source data directory.")
    ap.add_argument("--max-sample", type=int, default=cfg.get("maxSampleRows", 200))
    args = ap.parse_args()

    data_dir = dl.resolve_data_dir(args.data_dir)
    dl.log(f"scanning {data_dir}")

    if not data_dir.exists():
        dl.log(f"ERROR: data directory does not exist: {data_dir}")
        dl.log("Set it via --data-dir, the DRAFTLENS_DATA_DIR env var, or config/pipeline.config.json")
        return 2

    files = dl.discover_files(data_dir, cfg["fileTypes"])
    dl.log(f"found {len(files)} candidate files")

    profiles: list[dl.TableProfile] = []
    for sf in files:
        profiles.extend(profile_file(sf, args.max_sample))

    generated_at = datetime.now(timezone.utc).isoformat()
    inventory_obj = {
        "generatedAt": generated_at,
        "dataDir": str(data_dir),
        "tableCount": len(profiles),
        "tables": [vars(p) for p in profiles],
    }
    dl.write_json(dl.repo_path(cfg["processedDir"], "inventory.json"), inventory_obj)
    dl.write_text(dl.repo_path(cfg["docsDir"], "data_inventory.md"), render_markdown(data_dir, profiles, generated_at))

    dl.log(f"wrote {cfg['processedDir']}/inventory.json and {cfg['docsDir']}/data_inventory.md")
    n_warn = sum(len(p.warnings) for p in profiles)
    if n_warn:
        dl.log(f"{n_warn} warning(s) — see the inventory for details")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
